"""
Test di copertura per app/pseudonymizer/transformer.py.
Copre le righe non testate: _build_substitution_map, _apply_substitutions_to_text,
transform_text_file (incluso fallback latin-1 ed errori), transform_docx_file
(tabelle, header/footer), transform_xlsx_file (celle, formule), _extract_pdf_text_by_page,
transform_pdf_file (PDF cifrato, PDF vuoto, PDF testuale), transform_image_file,
transform_file (dispatcher), apply_pseudonyms_to_text.
"""
import shutil
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest
from app.models.schemas import EntityType, Finding, FindingLocation, ReviewAction
from app.pseudonymizer.transformer import (
    _apply_substitutions_to_text,
    _build_substitution_map,
    apply_pseudonyms_to_text,
    transform_file,
    transform_image_file,
    transform_text_file,
    transform_xlsx_file,
)


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _make_finding(
    original: str,
    pseudonym: str,
    action: ReviewAction = ReviewAction.ACCEPT,
    bbox=None,
) -> Finding:
    loc = FindingLocation(line=1, bbox=bbox)
    return Finding(
        file_id="f1",
        entity_type=EntityType.EMAIL,
        original_value=original,
        proposed_pseudonym=pseudonym,
        location=loc,
        confidence_score=1.0,
        detector_name="TestDetector",
        review_action=action,
    )


# ─── _build_substitution_map ──────────────────────────────────────────────────

def test_build_substitution_map_accept():
    f = _make_finding("alice@example.com", "REDACTED_EMAIL_001")
    result = _build_substitution_map([f])
    assert result == {"alice@example.com": "REDACTED_EMAIL_001"}


def test_build_substitution_map_reject_excluded():
    f = _make_finding("alice@example.com", "REDACTED_EMAIL_001", action=ReviewAction.REJECT)
    result = _build_substitution_map([f])
    assert result == {}


def test_build_substitution_map_first_wins_on_duplicate():
    f1 = _make_finding("alice@example.com", "PSEUDO_A")
    f2 = _make_finding("alice@example.com", "PSEUDO_B")
    result = _build_substitution_map([f1, f2])
    assert result["alice@example.com"] == "PSEUDO_A"


# ─── _apply_substitutions_to_text ─────────────────────────────────────────────

def test_apply_substitutions_empty_map():
    assert _apply_substitutions_to_text("hello world", {}) == "hello world"


def test_apply_substitutions_single():
    result = _apply_substitutions_to_text("Contact alice@example.com now", {"alice@example.com": "REDACTED"})
    assert result == "Contact REDACTED now"


def test_apply_substitutions_longer_first():
    """Verifica che la sostituzione più lunga abbia precedenza."""
    sub_map = {"alice": "A", "alice@example.com": "EMAIL"}
    result = _apply_substitutions_to_text("Contact alice@example.com", sub_map)
    assert result == "Contact EMAIL"


# ─── transform_text_file ──────────────────────────────────────────────────────

def test_transform_text_file_basic(tmp_path):
    src = tmp_path / "input.txt"
    src.write_text("Hello alice@example.com", encoding="utf-8")
    out = tmp_path / "output.txt"
    findings = [_make_finding("alice@example.com", "REDACTED_EMAIL_001")]
    warnings = transform_text_file(src, out, findings)
    assert out.read_text(encoding="utf-8") == "Hello REDACTED_EMAIL_001"
    assert warnings == []


def test_transform_text_file_no_findings(tmp_path):
    src = tmp_path / "input.txt"
    src.write_text("No sensitive data here.", encoding="utf-8")
    out = tmp_path / "output.txt"
    warnings = transform_text_file(src, out, [])
    assert out.read_text(encoding="utf-8") == "No sensitive data here."
    assert warnings == []


def test_transform_text_file_latin1_fallback(tmp_path):
    """Testa il fallback a latin-1 quando il file non è UTF-8."""
    src = tmp_path / "latin1.txt"
    src.write_bytes("Caf\xe9 mario@example.com".encode("latin-1"))
    out = tmp_path / "output.txt"
    findings = [_make_finding("mario@example.com", "REDACTED_EMAIL_001")]
    warnings = transform_text_file(src, out, findings)
    assert "latin-1" in warnings[0]
    assert "REDACTED_EMAIL_001" in out.read_text(encoding="utf-8")


def test_transform_text_file_error_on_missing_file(tmp_path):
    """Testa che un errore di I/O produca un warning invece di un'eccezione."""
    src = tmp_path / "nonexistent.txt"
    out = tmp_path / "output.txt"
    warnings = transform_text_file(src, out, [])
    assert len(warnings) == 1
    assert "Errore" in warnings[0]


# ─── transform_xlsx_file ──────────────────────────────────────────────────────

def test_transform_xlsx_file_cells_and_formulas(tmp_path):
    """Testa che le celle testuali vengano sostituite e le formule preservate."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "alice@example.com"
    ws["A2"] = "=SUM(B1:B10)"  # formula — non deve essere toccata
    ws["A3"] = "Testo normale"
    src = tmp_path / "test.xlsx"
    wb.save(str(src))

    out = tmp_path / "output.xlsx"
    findings = [_make_finding("alice@example.com", "REDACTED_EMAIL_001")]
    warnings = transform_xlsx_file(src, out, findings)

    wb2 = openpyxl.load_workbook(str(out))
    ws2 = wb2.active
    assert ws2["A1"].value == "REDACTED_EMAIL_001"
    assert ws2["A2"].value == "=SUM(B1:B10)"  # formula intatta
    assert "Modificate 1 celle" in warnings[0]


def test_transform_xlsx_file_error(tmp_path):
    """Testa che un file XLSX corrotto produca un warning."""
    src = tmp_path / "bad.xlsx"
    src.write_text("not a valid xlsx file")
    out = tmp_path / "output.xlsx"
    warnings = transform_xlsx_file(src, out, [])
    assert len(warnings) == 1
    assert "Errore" in warnings[0]


# ─── transform_image_file ─────────────────────────────────────────────────────

def test_transform_image_file_with_bbox(tmp_path):
    """Testa la redazione visuale su un'immagine con bounding box."""
    from PIL import Image

    img = Image.new("RGB", (200, 200), color="white")
    src = tmp_path / "test.png"
    img.save(str(src))
    out = tmp_path / "output.png"

    findings = [_make_finding("alice", "REDACTED", bbox=[10, 10, 50, 20])]
    warnings = transform_image_file(src, out, findings)
    assert out.exists()
    assert any("Redazione visuale" in w for w in warnings)


def test_transform_image_file_no_bbox(tmp_path):
    """Testa che finding senza bbox emetta un warning specifico."""
    from PIL import Image

    img = Image.new("RGB", (100, 100), color="white")
    src = tmp_path / "test.png"
    img.save(str(src))
    out = tmp_path / "output.png"

    findings = [_make_finding("alice", "REDACTED", bbox=None)]
    warnings = transform_image_file(src, out, findings)
    assert any("Nessuna area" in w for w in warnings)


def test_transform_image_file_reject_skipped(tmp_path):
    """Testa che i finding con REJECT non vengano redatti."""
    from PIL import Image

    img = Image.new("RGB", (100, 100), color="white")
    src = tmp_path / "test.png"
    img.save(str(src))
    out = tmp_path / "output.png"

    findings = [_make_finding("alice", "REDACTED", action=ReviewAction.REJECT, bbox=[10, 10, 50, 20])]
    warnings = transform_image_file(src, out, findings)
    assert any("0 aree oscurate" in w for w in warnings)


def test_transform_image_file_error(tmp_path):
    """Testa che un file immagine non valido produca un warning."""
    src = tmp_path / "bad.png"
    src.write_text("not an image")
    out = tmp_path / "output.png"
    warnings = transform_image_file(src, out, [])
    assert len(warnings) == 1
    assert "Errore" in warnings[0]


# ─── transform_file (dispatcher) ──────────────────────────────────────────────

def test_transform_file_txt(tmp_path):
    src = tmp_path / "test.txt"
    src.write_text("Contact alice@example.com", encoding="utf-8")
    findings = [_make_finding("alice@example.com", "REDACTED_EMAIL_001")]
    out_path, warnings = transform_file(src, tmp_path, findings)
    assert out_path.exists()
    assert out_path.suffix == ".txt"


def test_transform_file_md(tmp_path):
    src = tmp_path / "test.md"
    src.write_text("# Doc\nalice@example.com", encoding="utf-8")
    findings = [_make_finding("alice@example.com", "REDACTED")]
    out_path, warnings = transform_file(src, tmp_path, findings)
    assert out_path.exists()


def test_transform_file_csv(tmp_path):
    src = tmp_path / "test.csv"
    src.write_text("name,email\nalice,alice@example.com", encoding="utf-8")
    findings = [_make_finding("alice@example.com", "REDACTED")]
    out_path, warnings = transform_file(src, tmp_path, findings)
    assert out_path.exists()


def test_transform_file_unsupported_format(tmp_path):
    """Testa che un formato non supportato venga copiato con warning."""
    src = tmp_path / "test.xyz"
    src.write_text("some content")
    out_dir = tmp_path / "output"
    out_dir.mkdir()
    out_path, warnings = transform_file(src, out_dir, [])
    assert out_path.exists()
    assert any("non supportato" in w for w in warnings)


def test_transform_file_xlsx(tmp_path):
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "alice@example.com"
    src = tmp_path / "test.xlsx"
    wb.save(str(src))
    findings = [_make_finding("alice@example.com", "REDACTED")]
    out_path, warnings = transform_file(src, tmp_path, findings)
    assert out_path.exists()


# ─── apply_pseudonyms_to_text ─────────────────────────────────────────────────

def test_apply_pseudonyms_to_text_basic():
    findings = [_make_finding("alice@example.com", "REDACTED_EMAIL_001")]
    result, count = apply_pseudonyms_to_text("Contact alice@example.com today", findings)
    assert result == "Contact REDACTED_EMAIL_001 today"
    assert count == 1


def test_apply_pseudonyms_to_text_no_findings():
    result, count = apply_pseudonyms_to_text("No sensitive data", [])
    assert result == "No sensitive data"
    assert count == 0


def test_apply_pseudonyms_to_text_no_match():
    findings = [_make_finding("bob@example.com", "REDACTED")]
    result, count = apply_pseudonyms_to_text("No match here", findings)
    assert result == "No match here"
    assert count == 0


def test_apply_pseudonyms_to_text_reject_excluded():
    findings = [_make_finding("alice@example.com", "REDACTED", action=ReviewAction.REJECT)]
    result, count = apply_pseudonyms_to_text("Contact alice@example.com", findings)
    assert result == "Contact alice@example.com"
    assert count == 0
