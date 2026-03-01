"""
Parser per file .xlsx (Microsoft Excel).
Legge solo celle testuali; ignora formule e celle non-stringa.
"""
import logging
from pathlib import Path
from typing import List

from app.parsers.base import BaseParser, ParseResult, TextChunk
from app.core.exceptions import XlsxParsingError

logger = logging.getLogger(__name__)


class XlsxParser(BaseParser):
    """Parser per file Microsoft Excel (.xlsx)."""

    @property
    def supported_extensions(self) -> List[str]:
        return [".xlsx"]

    def parse(self, file_path: Path) -> ParseResult:
        result = ParseResult(file_path=file_path)
        formula_count = 0
        text_cell_count = 0

        try:
            import openpyxl
            try:
                # data_only=True legge i valori calcolati, non le formule.
                # Tuttavia, per identificare le formule, dobbiamo leggere anche con data_only=False.
                wb_formulas = openpyxl.load_workbook(str(file_path), data_only=False)
                wb_values = openpyxl.load_workbook(str(file_path), data_only=True)
            except Exception as load_err:
                raise XlsxParsingError(str(file_path), f"Failed to open/read XLSX: {load_err}") from load_err

            for sheet_name in wb_formulas.sheetnames:
                ws_formulas = wb_formulas[sheet_name]
                ws_values = wb_values[sheet_name]

                for row in ws_formulas.iter_rows():
                    for cell_formula in row:
                        cell_value = ws_values[cell_formula.coordinate]

                        # Controlla se la cella contiene una formula
                        if isinstance(cell_formula.value, str) and cell_formula.value.startswith("="):
                            formula_count += 1
                            # Aggiungi un chunk marcato come formula (non verrà processato)
                            result.chunks.append(
                                TextChunk(
                                    text=str(cell_formula.value),
                                    source_ref=f"foglio '{sheet_name}', cella {cell_formula.coordinate}",
                                    sheet_name=sheet_name,
                                    cell_ref=cell_formula.coordinate,
                                    is_formula=True,
                                )
                            )
                        elif isinstance(cell_value.value, str) and cell_value.value.strip():
                            # Cella con valore testuale
                            text_cell_count += 1
                            result.chunks.append(
                                TextChunk(
                                    text=cell_value.value,
                                    source_ref=f"foglio '{sheet_name}', cella {cell_formula.coordinate}",
                                    sheet_name=sheet_name,
                                    cell_ref=cell_formula.coordinate,
                                    is_formula=False,
                                )
                            )

            if formula_count > 0:
                result.warnings.append(
                    f"Trovate {formula_count} celle con formule: NON sono state modificate (come da policy MVP)."
                )
            result.warnings.append(
                f"Processate {text_cell_count} celle testuali su {text_cell_count + formula_count} celle totali analizzate."
            )

        except XlsxParsingError as e:
            result.success = False
            result.error_message = str(e)
            logger.warning("XLSX parsing error: %s", e)
        except Exception as e:
            result.success = False
            result.error_message = f"Errore durante il parsing del file XLSX: {e}"
            logger.error("Unexpected error in XLSX parser: %s", e)

        return result
